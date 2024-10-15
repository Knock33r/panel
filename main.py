import os
import subprocess
import threading
import queue
import openpyxl
import customtkinter as Ctk
from customtkinter import filedialog
import tkinter as tk
from tkinter import filedialog
from openpyxl.reader.excel import load_workbook
import downloadJar


# Variables globales
FichierSelec = None
entry = 'Nah'
process = None
console = None
chemin = 'Nah'
servOpen = None
fichierActuel = 'Aucun fichier selectionné'
serveurActuel = ''
serveurActuelId = None
pathDatas = r"C:\Users\cocop\PycharmProjects\panelSoftware\datas.xlsx"
dossierPrincipal = None

log_queue = queue.Queue()  # Queue pour gérer les logs

# Limite du nombre de lignes dans la console
MAX_LOG_LINES = 100


def closeWindow():
    window.destroy()


def demanderFichier():
    global FichierSelec
    global fichierActuel
    pathFinder = tk.Tk()
    pathFinder.withdraw()
    path = filedialog.askdirectory()
    fichierActuel = path
    FichierSelec.configure(text=fichierActuel)


def scanexel():
    datas = load_workbook(pathDatas)
    sheet = datas['Feuil1']
    compteur = 0

    for i in range(2, sheet.max_row + 1):
        if sheet[f'A{i}'].value is not None:
            compteur += 1
    return compteur


def fenetreAjoutServer():
    global FichierSelec, window, list

    def attribuervaleurs():
        global fichierActuel
        nom = NomServInput.get()
        path = fichierActuel
        port = PortServInput.get()
        query = QueryPortServInput.get()
        ram = RamMaxServInput.get()

        addServ(nom, path, port, query, ram)
        window.destroy()
        list.destroy()
        listServs()

    window = Ctk.CTk()
    window.geometry('800x800')

    NomServInput = Ctk.CTkEntry(window, placeholder_text='nom du serveur')
    FileServInput = Ctk.CTkButton(window, text='Chemin du fichier', command=demanderFichier)

    PortServInput = Ctk.CTkEntry(window, placeholder_text='port du serveur')
    QueryPortServInput = Ctk.CTkEntry(window, placeholder_text='25565')
    RamMaxServInput = Ctk.CTkEntry(window, placeholder_text='8')

    whitespace = Ctk.CTkLabel(window, text='')
    avancee = Ctk.CTkLabel(window, text='Options avancées. Laisser vide pour défaut.')
    NomServQ = Ctk.CTkLabel(window, text='Nom du serveur : ')
    FileServQ = Ctk.CTkLabel(window, text='Repertoire du serveur : ')

    FichierSelec = Ctk.CTkLabel(window, text=fichierActuel)

    PortServQ = Ctk.CTkLabel(window, text='Port du serveur : ')
    QueryPortServQ = Ctk.CTkLabel(window, text='Query port du serveur : ')
    RamMaxServQ = Ctk.CTkLabel(window, text='Ram maximale du serveur')

    valider = Ctk.CTkButton(window, text='Valider', command=attribuervaleurs)
    cancel = Ctk.CTkButton(window, text='Cancel', command=closeWindow)

    NomServQ.pack()
    NomServInput.pack()
    FileServQ.pack()
    FileServInput.pack()
    whitespace.pack()
    avancee.pack()
    PortServQ.pack()
    PortServInput.pack()
    QueryPortServQ.pack()
    QueryPortServInput.pack()
    RamMaxServQ.pack()
    RamMaxServInput.pack()
    valider.pack()
    cancel.pack()

    window.mainloop()


def addServ(nom, path, port, query, ram):
    datas = load_workbook(pathDatas)
    sheet = datas['Feuil1']
    nextLine = scanexel() + 2

    if port == null:
        port = 25565

    if query == null:
        query = port+1

    if ram == null:
        ram = 2

    sheet[f'A{nextLine}'] = str(nextLine)
    sheet[f'B{nextLine}'] = nom
    sheet[f'C{nextLine}'] = path
    sheet[f'D{nextLine}'] = port
    sheet[f'E{nextLine}'] = query
    sheet[f'F{nextLine}'] = ram

    datas.save(pathDatas)


def delServ():
    print('Nah')


def findPathById(id):
    datas = load_workbook(pathDatas)
    sheet = datas['Feuil1']
    path = sheet[f'C{id}'].value
    if path:
        print(f'chemin trouvé pour le serveur d id : {id} -> {path}.')
        return path
    else:
        print('Aucun serveur avec cet id n a été trouvé :/')


def ramUsageMaj():
    print('Nah')


def startServer():
    global servOpen, console, serveurActuelId, process
    file = r'C:\Users\cocop\Desktop\serv spi test'
    trouverFichier1(file)
    os.chdir(os.path.dirname(resultat))

    if serveurActuelId is None:
        print("Erreur : serveurActuelId n'est pas défini.")
        return

    process = execute_command(["cmd", "/c", resultat], console)


def trouver_fichier(path):
    for root, dirs, files in os.walk(path):
        for file in files:
            if file in ("run.bat", "start.bat"):
                return os.path.join(root, file)


def trouverFichier1(chemin):
    global resultat
    resultat = trouver_fichier(chemin)

    if resultat:
        print(f"Le fichier a été trouvé à : {resultat}")
    else:
        print("Aucun fichier run.bat ou start.bat n'a été trouvé.")


def send_command(event):
    global entry, process
    command = entry.get()
    if process is not None:
        process.stdin.write(command + "\n")
        process.stdin.flush()
    else:
        print("Erreur : le processus n'est pas initialisé.")
    entry.delete(0, Ctk.END)


def stop_serveur():
    global process
    if process is not None:
        process.stdin.write("stop" + "\n")
        process.stdin.flush()
    else:
        print("Erreur : le processus n'est pas initialisé.")


def execute_command(command, text_widget):
    global process
    process = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                               text=True)

    def stream_output():
        for line in iter(process.stdout.readline, ''):
            if line:
                log_queue.put(line)

    thread = threading.Thread(target=stream_output)
    thread.start()

    return process


def fenetreGestion(ligne):
    global console, process, entry, serveurActuelId
    serveurActuelId = ligne
    serveurActuel = findPathById(ligne)
    print(serveurActuel)

    datas = load_workbook(pathDatas)
    sheet = datas['Feuil1']
    gestion = Ctk.CTk()
    gestion.title('Gestion de serveur')
    gestion.after(0, lambda: gestion.state('zoomed'))

    def BtnMenu():
        gestion.destroy()
        listServs()

    menuBtn = Ctk.CTkButton(master=gestion, command=BtnMenu, text='MENU')
    menuBtn.pack()

    nomserveur = sheet[fr'B{ligne}'].value
    nomserv = Ctk.CTkLabel(gestion, text=nomserveur)
    nomserv.pack()
    console = Ctk.CTkTextbox(gestion, width=800, height=800)
    console.pack()

    entry = Ctk.CTkEntry(gestion)
    entry.pack()

    entry.bind("<Return>", send_command)
    start = Ctk.CTkButton(master=gestion, text='Start', command=startServer)
    stop = Ctk.CTkButton(master=gestion, text="Stop", command=stop_serveur)
    stop.pack()
    start.pack()

    gestion.after(100, lambda: start_log_thread(serveurActuelId, console))

    gestion.mainloop()


def start_log_thread(server_id, console_widget):
    log_file = f'server-id-{server_id}.chatlog'

    if not os.path.exists(log_file):
        with open(log_file, 'w') as f:
            pass

    with open(log_file, 'r') as f:
        for line in f:
            console_widget.insert(Ctk.END, line)
            console_widget.see(Ctk.END)

    def update_console():
        while True:
            try:
                line = log_queue.get_nowait()
                if line:
                    console_widget.insert(Ctk.END, line)
                    console_widget.see(Ctk.END)

                    # Limiter le nombre de lignes dans la console
                    if int(console_widget.index('end-1c').split('.')[0]) > MAX_LOG_LINES:
                        console_widget.delete(1.0, 2.0)

                    with open(log_file, 'a') as f:
                        f.write(line)
            except queue.Empty:
                break
        console_widget.after(100, update_console)

    console_widget.after(100, update_console)




def listServs():

    global list
    datas = load_workbook(pathDatas)
    sheet = datas['Feuil1']
    compteur = 0
    buttons = {}
    list = Ctk.CTk()
    list.title('Liste des serveurs')
    list.after(0, lambda: list.state('zoomed'))

    def actionBouton(i):
        global serveurActuelId
        list.destroy()
        serveurActuelId = i
        fenetreGestion(i)


    def fenetreCreationServer():
            global FichierSelec, window, list

            def attribuervaleurs():
                global fichierActuel
                nom = NomServInput.get()
                path = fichierActuel
                port = PortServInput.get()
                query = QueryPortServInput.get()
                ram = RamMaxServInput.get()

                addServ(nom, path, port, query, ram)
                window.destroy()
                list.destroy()
                listServs()

            window = Ctk.CTkToplevel()
            window.geometry('800x800')

            NomServInput = Ctk.CTkEntry(window, placeholder_text='nom du serveur')
            PortServInput = Ctk.CTkEntry(window, placeholder_text='port du serveur')
            QueryPortServInput = Ctk.CTkEntry(window, placeholder_text='25565')
            RamMaxServInput = Ctk.CTkEntry(window, placeholder_text='8')

            whitespace = Ctk.CTkLabel(window, text='')
            optAvancee = Ctk.CTkLabel(window, text='Options avancées. Laisser vide pour défaut.')
            NomServQ = Ctk.CTkLabel(window, text='Nom du serveur : ')

            PortServQ = Ctk.CTkLabel(window, text='Port du serveur : ')
            QueryPortServQ = Ctk.CTkLabel(window, text='Query port du serveur : ')
            RamMaxServQ = Ctk.CTkLabel(window, text='Ram maximale du serveur')

            valider = Ctk.CTkButton(window, text='Valider', command=attribuervaleurs)
            cancel = Ctk.CTkButton(window, text='Cancel', command=closeWindow)

            NomServQ.pack()
            NomServInput.pack()
            FileServQ.pack()
            FileServInput.pack()
            whitespace.pack()
            optAvancee.pack()
            PortServQ.pack()
            PortServInput.pack()
            QueryPortServQ.pack()
            QueryPortServInput.pack()
            RamMaxServQ.pack()
            RamMaxServInput.pack()
            valider.pack()
            cancel.pack()

            window.mainloop()

        
    btnCreerServer = Ctk.CTkButton(list, text='Créer un serveur', command=fenetreCreationServer)

    btnAddServer = Ctk.CTkButton(master=list, text='Lier un serveur', command=fenetreAjoutServer)
    espace=Ctk.CTkFrame(master=list, height=40, bg_color="black")

    btnCreerServer.pack()
    btnAddServer.pack()
    espace.pack()

    for i in range(2, sheet.max_row + 1):
        if sheet[f'A{i}'].value is not None:
            nom_serveur = sheet[f'B{i}'].value
            buttons[f'button{i}'] = Ctk.CTkButton(master=list, text=nom_serveur, width=750, height=90,command=lambda i=i: actionBouton(i))
            buttons[f'button{i}'].pack()
            compteur += 1

    list.mainloop()



listServs()